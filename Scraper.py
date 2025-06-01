import tkinter as tk
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import pickle
import time
import os
from screeninfo import get_monitors
from collections import deque
import threading
from openpyxl import Workbook, load_workbook

# Prompt for Excel file name
user_filename = input("Enter a name for the Excel log file (without extension): ").strip()
if not user_filename.endswith(".xlsx"):
    user_filename += ".xlsx"
excel_file = user_filename

# URLs and files
login_url = "https://xplay.gg/login"
dashboard_url = "https://xplay.gg/dashboard"
cookies_file = "cookies.pkl"

# Shared data
latest_xp_data = {"xp": None, "next_xp": None}
last_update_time = [time.time()]
last_change_time = [time.time()]
last_xp = [None]
xp_history = deque()
rate_history = {
    "2.5min": deque(),
    "5min": deque(),
    "12min": deque(),
}
last_gain = [0]

# For averages
last_2_5min_xp_values = []
exp_2_5min_xp_values = []
last_12min_xp_values = []
exp_5min_xp_values = []

# For delayed storing
pending_avg_update = [False]
avg_update_timer = [None]
xp_per_minute_global = [0]

RATE_WINDOW = 720  # 12 minutes in seconds

# Create Excel file and header if it doesn't exist
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Timestamp", "XP Gained (2.5min)", "Minutes", "Seconds"])
    wb.save(excel_file)

def update_xp_per_minute():
    total_xp = 0
    total_minutes = 0.0
    wb = load_workbook(excel_file)
    ws = wb.active
    # Skip the first two rows (header + first data row)
    for row in ws.iter_rows(min_row=3, values_only=True):
        xp, minutes, seconds = row[1], row[2], row[3]
        if xp is not None and minutes is not None and seconds is not None:
            total_xp += xp
            total_minutes += minutes + (seconds / 60.0)
    wb.close()
    if total_minutes > 0:
        xp_per_minute_global[0] = int(total_xp / total_minutes)

def log_to_excel(timestamp, xp_gain, elapsed_sec):
    minutes = int(elapsed_sec) // 60
    seconds = int(elapsed_sec) % 60
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
        ws.append([
            time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(timestamp)),
            xp_gain,
            minutes,
            seconds
        ])
        wb.save(excel_file)
        update_xp_per_minute()

def get_xp_with_cookies():
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    try:
        if os.path.exists(cookies_file):
            driver.get(login_url)
            cookies = pickle.load(open(cookies_file, "rb"))
            for cookie in cookies:
                driver.add_cookie(cookie)
            driver.get(dashboard_url)
            time.sleep(5)
            xp_element = driver.find_element(By.CLASS_NAME, 'sc-604d7992-5.jLFqvp')
            xp_value = xp_element.text.strip()
            current_xp, next_level_xp = xp_value.split(" / ")
            current_xp = int(current_xp.replace(",", "").strip())
            next_level_xp = int(next_level_xp.replace(",", "").replace(" xp", "").strip())
            return current_xp, next_level_xp
    except Exception as e:
        print(f"Error fetching XP: {e}")
    finally:
        driver.quit()
    return None

def schedule_avg_storage(current_xp):
    def store_avg_values():
        if not pending_avg_update[0]:
            return
        now = time.time()

        xp_2_5min = current_xp - xp_at(150)
        xp_5min = current_xp - xp_at(300)
        xp_12min = current_xp - xp_at(720)

        if xp_2_5min > 0:
            last_2_5min_xp_values.append(xp_2_5min)
            exp_2_5min_xp_values.append(int(xp_2_5min * 4.8))

        if xp_12min > 0:
            last_12min_xp_values.append(xp_12min)

        if xp_5min > 0:
            exp_5min_xp_values.append(int(xp_5min * 2.4))

        pending_avg_update[0] = False

    if avg_update_timer[0]:
        avg_update_timer[0].cancel()

    pending_avg_update[0] = True
    timer = threading.Timer(60, store_avg_values)
    avg_update_timer[0] = timer
    timer.start()

def fetch_xp_loop():
    while True:
        result = get_xp_with_cookies()
        if result:
            current_xp, next_level_xp = result
            now = time.time()
            prev_xp = latest_xp_data["xp"]
            prev_time = last_update_time[0]

            if prev_xp is not None and current_xp > prev_xp:
                elapsed = now - prev_time
                gain = current_xp - prev_xp
                last_gain[0] = gain

                schedule_avg_storage(current_xp)

                xp_2_5min = current_xp - xp_at(150)
                if xp_2_5min > 0:
                    time_since_last_change = now - last_change_time[0]
                    log_to_excel(now, xp_2_5min, time_since_last_change)
                    last_change_time[0] = now

                if elapsed > 0 and gain > 0:
                    rate = gain / elapsed
                    rate_history["2.5min"].append((now, rate, elapsed))
                    rate_history["5min"].append((now, rate, elapsed))
                    rate_history["12min"].append((now, rate, elapsed))

                last_update_time[0] = now

            last_xp[0] = current_xp
            latest_xp_data["xp"] = current_xp
            latest_xp_data["next_xp"] = next_level_xp

            xp_history.append((now, current_xp))
            while xp_history and now - xp_history[0][0] > RATE_WINDOW:
                xp_history.popleft()

        time.sleep(10)

def xp_at(seconds_ago):
    target = time.time() - seconds_ago
    for t, xp in reversed(xp_history):
        if t <= target:
            return xp
    return xp_history[0][1] if xp_history else latest_xp_data["xp"]

def create_overlay():
    root = tk.Tk()
    root.title("XP Overlay")

    monitors = get_monitors()
    if len(monitors) > 1:
        second_monitor = monitors[1]
        window_position = f"+{-second_monitor.width // 2 + 660}+{second_monitor.height // 2 + 345}"
    else:
        window_position = "+10+10"

    root.overrideredirect(True)
    root.attributes('-topmost', True)
    root.geometry(f"300x260{window_position}")

    label = tk.Label(root, text="Loading...", font=("Arial", 15), fg="white", bg="black", anchor="n", justify="center", pady=5)
    label.pack(fill=tk.BOTH, expand=True)

    def calculate_avg(values):
        return sum(values) // len(values) if values else 0

    def update_gui():
        now = time.time()
        current_xp = latest_xp_data["xp"]
        next_level_xp = latest_xp_data["next_xp"]

        if current_xp and next_level_xp:
            xp_needed = next_level_xp - current_xp
            xp_text = f"{current_xp} / {next_level_xp} ({xp_needed})"

            xp_12min = current_xp - xp_at(720)
            xp_2_5min = current_xp - xp_at(150)
            xp_5min = current_xp - xp_at(300)

            exp_2_5min = int(xp_2_5min * 4.8)
            exp_5min = int(xp_5min * 2.4)

            avg_real_2_5 = calculate_avg(last_2_5min_xp_values)
            avg_expected_2_5 = calculate_avg(exp_2_5min_xp_values)
            avg_real_12 = calculate_avg(last_12min_xp_values)
            avg_expected_5 = calculate_avg(exp_5min_xp_values)

            elapsed = int(now - last_update_time[0])
            minutes = elapsed // 60
            seconds = elapsed % 60

            label.config(
                text=f"{xp_text}\n"
                     f"Last 12min: {xp_12min} ({avg_real_12})\n"
                     f"Last 2.5min: {xp_2_5min} ({avg_real_2_5})\n"
                     f"Exp 5min: {exp_5min} ({avg_expected_5})\n"
                     f"Exp 2.5min: {exp_2_5min} ({avg_expected_2_5})\n"
                     f"Last Gain XP: {last_gain[0]}\n"
                     f"Last Update: {minutes}m {seconds}s\n"
                     f"XP PER MIN: {xp_per_minute_global[0]}"
            )

        root.after(1000, update_gui)

    update_gui()
    root.mainloop()

def save_cookies_manually():
    options = Options()
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    try:
        driver.get(login_url)
        print("Please log in manually...")
        input("Press Enter once logged in...")
        cookies = driver.get_cookies()
        with open(cookies_file, "wb") as f:
            pickle.dump(cookies, f)
        print("Cookies saved.")
    finally:
        driver.quit()

if not os.path.exists(cookies_file):
    save_cookies_manually()

threading.Thread(target=fetch_xp_loop, daemon=True).start()
create_overlay()
